#coding:utf-8

"""
Version: 0.95
Date: 2020-12-10
Author: Kyrie403
link: https://github.com/kyrie403
Usage: python3 GobyParser.py filename.xlsx
       python3 GobyParser.py -d dir_name
"""

import os
import sys
import json
import pymysql
import openpyxl

db = pymysql.connect(host='127.0.0.1',
                     user='root',
                     password='root',
                     port=3306,
                     database='goby',
                     charset='utf8')
cursor = db.cursor()
db.autocommit(True)

def get_file_list(dir_name):
    file_list = []
    for root, dirs, files in os.walk(dir_name, topdown=True):
        file_list.extend([os.path.join(root, file_name) for file_name in files if os.path.splitext(os.path.join(root, file_name))[1] == '.xlsx'])
    return file_list

def get_data_from_excel(path):
    if os.path.exists(path):
        asset_column_name = {1: "IP", 2: "Port", 3: "Protocol", 4: "Mac", 5: "Host", 6: "Application Layer",
                             7: "Support Layer", 8: "Service Layer", 9: "System Layer", 10: "Hardware Layer"}
        vuln_column_name = {1: "filename", 2: "level", 3: "hostinfo", 4: "vulurl", 5: "keymemo"}
        data_list = {}
        data_info = {}
        ip = ""
        host_info = ""
        excel = openpyxl.load_workbook(path)
        sheet = excel.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        column_name = asset_column_name if max_column == 10 else vuln_column_name
        for i in range(2, max_row+1):
            for j in range(1, max_column+1):
                if max_column == 10 and j == 1:
                    ip = sheet.cell(row=i, column=j).value
                if max_column != 10 and j == 3:
                    host_info = sheet.cell(row=i, column=j).value
                cell = sheet.cell(row=i, column=j)
                data_info[column_name[j]] = cell.value
            if max_column == 10:
                data_list[ip] = data_info
            else:
                data_list[host_info] = data_info
            data_info = {}
        return data_list
    else:
        print("[ERROR]: File does not exist")
        exit(0)

def create_table():
    sql = """CREATE TABLE IF NOT EXISTS `asset` (
      `id` int(11) NOT NULL AUTO_INCREMENT,
      `ip` VARCHAR(255) NOT NULL,
      `Port` TEXT(2555) DEFAULT NULL,
      `Protocol` TEXT(2555) DEFAULT NULL,
      `Mac` VARCHAR(255) DEFAULT NULL,
      `Host` VARCHAR(255) DEFAULT NULL,
      `Application Layer` TEXT(2555) DEFAULT NULL,
      `Support Layer` TEXT(2555) DEFAULT NULL,
      `Service Layer` TEXT(2555) DEFAULT NULL,
      `System Layer` TEXT(2555) DEFAULT NULL,
      `Hardware Layer` TEXT(2555) DEFAULT NULL,
      PRIMARY KEY (`ip`),
      KEY `id` (`id`)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8;""" + """CREATE TABLE IF NOT EXISTS `vuln` (
      `id` int(11) NOT NULL AUTO_INCREMENT,
      `hostinfo` VARCHAR(255) NOT NULL,
      `filename` VARCHAR(255) DEFAULT NULL,
      `level` VARCHAR(255) DEFAULT NULL,
      `vulurl` VARCHAR(255) DEFAULT NULL,
      `keymemo` VARCHAR(255) DEFAULT NULL,
      PRIMARY KEY (`hostinfo`),
      KEY `id` (`id`)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8;"""
    cursor.execute(sql)


def update_data(data_list):
    for key, value in data_list.items():
        sql = """UPDATE `asset` set `Port`="{Port}", `Protocol`="{Protocol}", `Mac`="{Mac}", `Host`="{Host}", `Application Layer`="{Application_Layer}", `Support Layer`="{Support_Layer}", `Service Layer`="{Service_Layer}", `System Layer`="{System_Layer}", `Hardware Layer`="{Hardware_Layer}" WHERE `ip`="{ip}";""".format(ip=key, Port=value['Port'], Protocol=value['Protocol'], Mac=value['Mac'], Host=value['Host'], Application_Layer=value['Application Layer'], Support_Layer=value['Support Layer'], Service_Layer=value['Service Layer'], System_Layer=value['System Layer'], Hardware_Layer=value['Hardware Layer']) \
        if 'hostinfo' not in value.keys() else\
        """UPDATE `vuln` set `filename`="{filename}", `level`="{level}", `vulurl`="{vulurl}", `keymemo`="{keymemo}" WHERE `hostinfo`="{hostinfo}";""".format(hostinfo=key, filename=value['filename'], level=value['level'], vulurl=value['vulurl'], keymemo=value['keymemo'])
        try:
            cursor.execute(sql)
        except pymysql.err.ProgrammingError:
            db.rollback()
            print(sql)

def insert_data(data_list):
    for key, value in data_list.items():
        sql = """INSERT INTO `asset` (`ip`, `Port`, `Protocol`, `Mac`, `Host`, 
    `Application Layer`, `Support Layer`, `Service Layer`, `System Layer`, `Hardware Layer`) VALUES 
    ("{ip}", "{Port}", "{Protocol}", "{Mac}", "{Host}", 
    "{Application_Layer}", "{Support_Layer}", "{Service_Layer}", "{System_Layer}","{Hardware_Layer}");""".format(ip=key, Port=value['Port'], Protocol=value['Protocol'], Mac=value['Mac'], Host=value['Host'], Application_Layer=value['Application Layer'], Support_Layer=value['Support Layer'], Service_Layer=value['Service Layer'], System_Layer=value['System Layer'], Hardware_Layer=value['Hardware Layer']) \
            if 'hostinfo' not in value.keys() else\
            """INSERT INTO `vuln` (`hostinfo`, `filename`, `level`, `vulurl`, `keymemo`) VALUES 
    ("{hostinfo}", "{filename}", "{level}", "{vulurl}", "{keymemo}");""".format(hostinfo=key, filename=value['filename'], level=value['level'], vulurl=value['vulurl'], keymemo=value['keymemo'])
        try:
            cursor.execute(sql)
        except pymysql.err.ProgrammingError:
            db.rollback()
            print(sql)
        except pymysql.err.IntegrityError:
            data_list = {key: value}
            update_data(data_list)

def task():
    if len(sys.argv) >= 2:
        print("[info]: Start")
        file_list = get_file_list(sys.argv[2]) if sys.argv[1] == '-d' and len(sys.argv) >= 3 else [sys.argv[1]]
        create_table()
        for file in file_list:
            data_list = get_data_from_excel(file)
            output_file_name = "{}.json".format(file)
            f = open(output_file_name, "w", encoding="utf-8")
            f.write(json.dumps(data_list))
            f.close()
            insert_data(data_list)
            print("[info]: {} import complete".format(file))
        db.close()
        print("[info]: Done")
    else:
        print("[Usage]: python3 GobyParser.py filename.xlsx \npython3 GobyParser.py -d dir_name")

if __name__ == '__main__':
    task()


